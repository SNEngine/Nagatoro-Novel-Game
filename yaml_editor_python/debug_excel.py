import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from openpyxl import load_workbook

# Load the Excel file and check the worksheet structure
workbook = load_workbook('test_export_full.xlsx')

for sheet_name in workbook.sheetnames:
    print(f"Sheet: {sheet_name}")
    worksheet = workbook[sheet_name]
    
    # Print first few rows
    rows = list(worksheet.iter_rows(values_only=True))
    for i, row in enumerate(rows[:5]):  # Print first 5 rows
        print(f"  Row {i}: {row}")
    print()