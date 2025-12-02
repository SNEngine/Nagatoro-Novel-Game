"""
Module for handling Excel import and export functionality in the YAML editor.
This implementation now handles the complete language folder structure as required.
"""
import os
import yaml
from typing import Dict, Any, List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QProgressDialog
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor


class ExcelHandler:
    """
    Handles import/export operations between Excel files and the complete language folder structure.
    """

    def __init__(self, parent_window=None):
        self.parent_window = parent_window
        self.styles = getattr(parent_window, 'STYLES', {}) if parent_window else {}

    def import_from_excel(self, file_path: str, target_folder: str) -> bool:
        """
        Import data from Excel file and restore the complete language folder structure.

        Args:
            file_path: Path to the Excel file containing exported language data
            target_folder: Target folder to restore the language structure to

        Returns:
            True if successful, False otherwise
        """
        try:
            # Load the Excel workbook
            workbook = load_workbook(file_path)

            success_count = 0
            total_count = 0

            # Process each worksheet as a separate YAML file
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Skip empty sheets
                if not self._has_data(worksheet):
                    continue

                total_count += 1

                # Convert worksheet to YAML data
                yaml_data = self._worksheet_to_yaml(worksheet)

                # Determine filename from worksheet name
                # If sheet name corresponds to a standard file, use that name
                # Otherwise, create a generic filename
                # We want to avoid adding unnecessary language code prefixes
                # but also handle cases where the sheet name might already contain them
                sheet_name_lower = sheet_name.lower()

                # Check if this is a standard filename like 'ui', 'characters', etc.
                if sheet_name_lower in ['metadata', 'characters', 'ui', 'terms']:
                    filename = f"{sheet_name_lower}.yaml"
                else:
                    # For other names, remove potential language code prefixes if they're followed by common standard names
                    # e.g., if we have 'en_ui' where 'ui' is a standard name, use just 'ui'
                    parts = sheet_name_lower.split('_')
                    if len(parts) > 1:
                        # Check if the last part is a standard name
                        if parts[-1] in ['metadata', 'characters', 'ui', 'terms']:
                            # Use just the last part as the filename (e.g., 'ui' from 'en_ui')
                            filename = f"{parts[-1]}.yaml"
                        else:
                            # Use the full name as a filename (e.g., 'custom_file' from 'en_custom_file')
                            filename = f"{sheet_name_lower.replace(' ', '_')}.yaml"
                    else:
                        # For single-part names, just sanitize and use
                        filename = f"{sheet_name_lower.replace(' ', '_')}.yaml"

                # Write the YAML data to a file in the target folder
                # Handle potential subdirectory paths in the filename
                # For example, sheet name 'dialogues_dialogue_2' should create 'dialogues/dialogue_2.yaml'
                import os

                # Determine the actual file path based on the computed filename
                if '_' in filename and not any(filename.endswith(f"{std_name}.yaml") for std_name in ['metadata', 'characters', 'ui', 'terms']):
                    parts = filename.split('_')
                    if len(parts) >= 2 and len(parts[0]) > 0:  # Make sure there's a potential subfolder name
                        # Check if the first part was likely a subdirectory in the original structure
                        subfolder = parts[0].replace('.yaml', '')  # Potential subfolder like 'dialogues'

                        # Common subdirectories that are likely in original structure
                        likely_subdirs = ['dialogues', 'levels', 'screens', 'scenes', 'chapters', 'quests', 'items', 'characters_data']

                        if subfolder in likely_subdirs:
                            # Original was in a subdirectory, recreate the path
                            actual_filename_without_yaml = '_'.join(parts[1:])  # e.g., 'dialogue_2' (without '.yaml')
                            if actual_filename_without_yaml.endswith('.yaml'):
                                actual_filename = actual_filename_without_yaml
                            else:
                                actual_filename = actual_filename_without_yaml + '.yaml'  # e.g., 'dialogue_2.yaml'
                            subfolder_path = os.path.join(target_folder, subfolder)
                            os.makedirs(subfolder_path, exist_ok=True)
                            file_path = os.path.join(subfolder_path, actual_filename)
                        else:
                            # Just a regular filename with underscores, save in main directory
                            file_path = os.path.join(target_folder, filename)
                else:
                    # Regular filename, save in main directory
                    file_path = os.path.join(target_folder, filename)

                # Ensure the directory exists
                directory_path = os.path.dirname(file_path)
                # Create the target folder if it doesn't exist (for cases where it's needed)
                os.makedirs(target_folder, exist_ok=True)
                # Create any subdirectory if needed
                if directory_path and directory_path != target_folder:
                    os.makedirs(directory_path, exist_ok=True)

                try:
                    # The yaml_data is the actual parsed YAML structure, save it directly
                    with open(file_path, 'w', encoding='utf-8') as f:
                        yaml.dump(yaml_data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)

                    success_count += 1
                except Exception as e:
                    print(f"Error writing file {filename}: {str(e)}")
                    continue

            if self.parent_window:
                if success_count > 0:
                    color = QColor(self.styles.get('DarkTheme', {}).get('NotificationSuccess', '#6BA878'))
                    self.parent_window.show_notification(
                        f"Successfully imported {success_count}/{total_count} files from Excel", color
                    )
                    return True
                else:
                    color = QColor(self.styles.get('DarkTheme', {}).get('NotificationError', '#D9685A'))
                    self.parent_window.show_notification("No files were imported from Excel", color)
                    return False

            return success_count > 0

        except Exception as e:
            error_msg = f"Error importing Excel file: {str(e)}"
            print(error_msg)
            if self.parent_window:
                color = QColor(self.styles.get('DarkTheme', {}).get('NotificationError', '#D9685A'))
                self.parent_window.show_notification(error_msg, color)
            return False

    def _has_data(self, worksheet: Worksheet) -> bool:
        """
        Check if worksheet contains actual data (not just empty cells).
        """
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    return True
        return False

    def _worksheet_to_yaml(self, worksheet: Worksheet) -> Dict[str, Any]:
        """
        Convert a single Excel worksheet to YAML data structure.

        Args:
            worksheet: OpenPyXL worksheet object

        Returns:
            Dictionary representing the YAML structure
        """
        # Get all rows from the worksheet
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            return {}

        # Check if this is our new improved format with "Key", "Value", "Original YAML Type" in the header
        first_row = rows[0]
        if (len(first_row) >= 3 and
            str(first_row[0]).strip() == "Key" and
            str(first_row[1]).strip() == "Value" and
            str(first_row[2]).strip() == "Original YAML Type"):

            # Look for the row with "RAW_YAML_DATA" marker which contains the backup YAML
            for row in rows[1:]:  # Skip header row
                if len(row) >= 3 and str(row[0]).strip() == "RAW_YAML_DATA":
                    yaml_content = row[1]  # Second column contains the YAML string
                    try:
                        # Parse the YAML content back to Python data structure
                        parsed_data = yaml.safe_load(yaml_content) if yaml_content is not None else {}
                        # Return the parsed data directly (not wrapped in another dict)
                        return parsed_data if isinstance(parsed_data, dict) else {'data': parsed_data}
                    except yaml.YAMLError as e:
                        print(f"Error parsing YAML from Excel: {e}")
                        # If YAML parsing fails, try JSON parsing as fallback
                        try:
                            import json
                            return json.loads(yaml_content) if yaml_content is not None else {}
                        except:
                            print("Fallback parsing also failed")
                            return {}

            # If no RAW_YAML_DATA found, process rows as key-value pairs
            result = {}
            for row in rows[1:]:  # Skip header row
                if len(row) >= 2:
                    key = row[0]
                    value = row[1]
                    if key is not None:
                        # Try to parse the value back to its original type if possible
                        if value is not None:
                            result[str(key)] = self._parse_yaml_if_needed(value)
                        else:
                            result[str(key)] = None
            return result

        # Check if this is the old special format with "Type" and "Content" in the header
        elif (len(first_row) >= 2 and
              str(first_row[0]).strip() == "Type" and
              str(first_row[1]).strip() == "Content"):
            # This is our old special format - look for the row with "YAML_DATA" marker
            for row in rows[1:]:  # Skip header row
                if len(row) >= 2 and str(row[0]).strip() == "YAML_DATA":
                    yaml_content = row[1]  # Second column contains the YAML string
                    try:
                        # Parse the YAML content back to Python data structure
                        parsed_data = yaml.safe_load(yaml_content) if yaml_content is not None else {}
                        # Return the parsed data directly (not wrapped in another dict)
                        return parsed_data if isinstance(parsed_data, dict) else {'data': parsed_data}
                    except yaml.YAMLError as e:
                        print(f"Error parsing YAML from Excel: {e}")
                        # If YAML parsing fails, try JSON parsing as fallback
                        try:
                            import json
                            return json.loads(yaml_content) if yaml_content is not None else {}
                        except:
                            print("Fallback parsing also failed")
                            return {}

        # Otherwise, use the old processing method for compatibility
        if self._looks_like_headers(first_row):
            # Process with headers
            return self._process_rows_with_headers(rows[1:], first_row)
        else:
            # Process without headers - assume simple key-value pairs
            return self._process_rows_without_headers(rows)

    def _looks_like_headers(self, row) -> bool:
        """
        Determine if the row looks like headers based on content.
        """
        if not row or len(row) == 0:
            return False

        # Check if first several cells contain text that looks like headers
        header_count = 0
        for i, cell in enumerate(row):
            if cell is not None:
                cell_str = str(cell).strip()
                if (cell_str and
                    not cell_str.startswith(" ") and
                    not cell_str.endswith(":") and
                    not self._is_typical_data_value(cell_str)):
                    header_count += 1
                    # If we have at least 1 potential header, consider it
                    if header_count >= 1:
                        return True
            # Only check first few columns
            if i >= 5:
                break

        return header_count > 0

    def _is_typical_data_value(self, value: str) -> bool:
        """
        Check if value is a typical data value (number, boolean, etc.) rather than a header.
        """
        value_lower = value.lower()
        # Check for common boolean values
        if value_lower in ["true", "false", "yes", "no", "1", "0", "null", "none"]:
            return True

        # Check if it's a number (int or float)
        try:
            float(value)
            return True
        except ValueError:
            pass

        return False

    def _process_rows_with_headers(self, data_rows: List[tuple], headers: tuple) -> Dict[str, Any]:
        """
        Process rows that have headers in the first row.
        """
        result = {}

        # Create a list of header names from the first row
        header_names = []
        for i, header in enumerate(headers):
            if header is not None:
                header_names.append(str(header).strip())
            else:
                header_names.append(f"Column_{i}")

        # Check if this is a special format (like "Category", "Item_Index", "Value")
        if len(header_names) == 3 and header_names[0] == "Category" and header_names[1] == "Item_Index" and header_names[2] == "Value":
            # This is the format used for structured data with lists of dicts
            result = self._process_structured_format(data_rows, header_names)
        elif len(header_names) >= 2 and header_names[0] == "ID":
            # This is the format with ID and other properties
            result = self._process_id_property_format(data_rows, header_names)
        else:
            # Standard format processing
            for row in data_rows:
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue  # Skip empty rows

                # Use the first column as the key for the outer dictionary
                key = str(row[0]) if row and len(row) > 0 and row[0] is not None else f"row_{len(result)+1}"

                # Create a dictionary of key-value pairs from the rest of the row
                row_data = {}
                for i, header in enumerate(header_names[1:]):  # Skip the first header since it's used as key
                    if i < len(row) - 1:  # Make sure we don't go out of bounds
                        value = row[i + 1]  # Skip first column since it's used as the key
                        if value is None:
                            value = ""
                        # If value looks like a YAML structure, try to parse it
                        row_data[header] = self._parse_yaml_if_needed(value)

                result[key] = row_data

        return result

    def _process_structured_format(self, data_rows: List[tuple], header_names: List[str]) -> Dict[str, Any]:
        """
        Process structured format with Category, Item_Index, Value columns.
        """
        result = {}

        for row in data_rows:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue  # Skip empty rows

            if len(row) >= 3:
                category = str(row[0]) if row[0] is not None else ""
                item_index = str(row[1]) if row[1] is not None else ""
                value = row[2] if row[2] is not None else ""

                # Try to parse if it looks like YAML
                value = self._parse_yaml_if_needed(value)

                if category and category not in result:
                    result[category] = []

                # If item_index looks like "item_N" (from a list), add to list
                if item_index.startswith("item_"):
                    # Add to the list for this category
                    result[category].append(value)
                elif item_index.strip() != "":
                    # If there's a specific item_index, add as dict entry
                    if not isinstance(result[category], dict):
                        result[category] = {}  # Convert to dict if not already
                    result[category][item_index] = value
                else:
                    # If no item_index, add as a simple value or append to list
                    if not isinstance(result[category], list):
                        result[category] = []  # Convert to list if not already
                    result[category].append(value)

        return result

    def _process_id_property_format(self, data_rows: List[tuple], header_names: List[str]) -> Dict[str, Any]:
        """
        Process format with ID column and property columns.
        """
        result = {}

        for row in data_rows:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue  # Skip empty rows

            # First column is the ID
            if len(row) > 0:
                id_val = str(row[0]) if row[0] is not None else f"item_{len(result)+1}"

                # Process remaining columns as properties
                row_data = {}
                for i, header in enumerate(header_names[1:]):  # Skip ID column
                    if i < len(row) - 1:  # Make sure we don't go out of bounds
                        value = row[i + 1]  # Skip first column (ID)
                        if value is None:
                            value = ""
                        # Try to parse if it looks like YAML
                        value = self._parse_yaml_if_needed(value)
                        row_data[header] = value

                result[id_val] = row_data

        return result

    def _parse_yaml_if_needed(self, value: Any) -> Any:
        """
        Try to parse a value as YAML if it looks like it contains structured data.
        """
        if not isinstance(value, str):
            return value

        value_str = value.strip()

        # Check if value looks like YAML/JSON structure
        if (value_str.startswith('{') and value_str.endswith('}')) or \
           (value_str.startswith('[') and value_str.endswith(']')) or \
           (':' in value_str and ('\n' in value_str or '\r' in value_str)):
            try:
                # Try to parse as YAML
                parsed = yaml.safe_load(value_str)
                return parsed if parsed is not None else value
            except:
                # If parsing fails, return original value
                pass

        return value

    def _process_rows_without_headers(self, rows: List[tuple]) -> Dict[str, Any]:
        """
        Process rows without headers, assuming they represent simple key-value pairs.
        """
        result = {}

        for row in rows:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue  # Skip empty rows

            # If we have at least 2 columns, treat as key-value pairs
            if len(row) >= 2 and row[0] is not None:
                key = str(row[0]).strip()
                value = row[1] if len(row) > 1 and row[1] is not None else ""
                result[key] = value
            # If we have only 1 column, use its index as a key
            elif len(row) >= 1 and row[0] is not None:
                key = f"item_{len(result)+1}"
                result[key] = row[0]

        return result

    def export_to_excel(self, source_folder: str, file_path: str) -> bool:
        """
        Export the complete language folder structure to an Excel file.
        Each YAML file becomes a separate worksheet.

        Args:
            source_folder: Source folder containing the language structure to export
            file_path: Path to save the Excel file

        Returns:
            True if successful, False otherwise
        """
        try:
            # Create a new workbook
            workbook = Workbook()

            # Remove the default empty sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            success_count = 0
            total_count = 0

            # Find all YAML files in the source folder (including subdirectories)
            yaml_files = self._find_yaml_files(source_folder)

            for yaml_file in yaml_files:
                total_count += 1

                try:
                    # Load the YAML file content
                    with open(yaml_file, 'r', encoding='utf-8') as f:
                        yaml_data = yaml.safe_load(f)

                    if yaml_data is None:
                        yaml_data = {}

                    # Create a worksheet name from the file path
                    # Remove the source folder path and the .yaml extension
                    rel_path = os.path.relpath(yaml_file, source_folder)

                    # Remove language code prefix from the folder name if present
                    # Extract folder name from source_folder path
                    folder_name = os.path.basename(source_folder.rstrip(os.sep))

                    # If folder_name looks like a language code (contains common language codes),
                    # remove it from the rel_path to avoid prefixes like 'en_ui'
                    # Common language codes: en, ru, es, fr, de, ja, ko, zh, etc.
                    lang_codes = ['en', 'ru', 'es', 'fr', 'de', 'ja', 'ko', 'zh', 'it', 'pt', 'nl', 'sv', 'no', 'da', 'fi', 'pl', 'cs', 'sk', 'hu', 'ro', 'bg', 'hr', 'sr', 'sl', 'et', 'lv', 'lt', 'uk', 'be', 'tr', 'ar', 'he', 'fa', 'ur', 'hi', 'bn', 'ta', 'te', 'ml', 'kn', 'mr', 'gu', 'pa', 'th', 'vi', 'id', 'ms', 'tl', 'km', 'my', 'lo', 'ka', 'el', 'is', 'ga', 'cy']

                    # Check if folder_name looks like a language code pattern
                    sheet_name = rel_path
                    for lang_code in lang_codes:
                        # Check for patterns like 'en', 'en_test', 'en-US', etc.
                        if (folder_name.lower() == lang_code or  # Exact match like 'en'
                            folder_name.lower().startswith(lang_code + '_') or  # Like 'en_test'
                            folder_name.lower().startswith(lang_code + '-')):  # Like 'en-US'

                            # If the folder name is exactly the language code, remove it from rel_path
                            if folder_name.lower() == lang_code:
                                # For example, if folder_name is 'en' and rel_path is 'ui.yaml' (for a subfolder like en/ui.yaml),
                                # the path might be different, so we might need to check if it starts with the lang_code
                                # Actually, let's think differently: if we export from 'en' folder,
                                # then the files inside it like 'en/ui.yaml' would have rel_path 'ui.yaml'
                                # since we're doing relpath(ui.yaml, en_folder)
                                # So we don't need to remove anything in this case, just keep the original rel_path
                                break
                            elif folder_name.lower().startswith(lang_code + '_'):
                                # If folder name is like 'en_test' and rel_path is 'en_test/something.yaml'
                                if rel_path.startswith(folder_name + os.sep) or rel_path.startswith(folder_name + '/'):
                                    sheet_name = rel_path[len(folder_name) + 1:]  # +1 for the separator
                                    break
                                # If folder name is like 'en_test' but file is just inside like 'ui.yaml'
                                # then rel_path would just be 'ui.yaml', so no prefix to remove
                                # That's already correct
                                break
                            elif folder_name.lower().startswith(lang_code + '-'):
                                # Similar for codes with dashes like 'en-US'
                                if rel_path.startswith(folder_name + os.sep) or rel_path.startswith(folder_name + '/'):
                                    sheet_name = rel_path[len(folder_name) + 1:]  # +1 for the separator
                                    break
                                break
                            break

                    # Replace directory separators with underscores and remove extensions
                    sheet_name = sheet_name.replace(os.sep, '_').replace('/', '_').replace('.yaml', '').replace('.yml', '')
                    # Limit length and remove invalid characters for Excel sheet names
                    sheet_name = self._sanitize_sheet_name(sheet_name)

                    # Create a new worksheet
                    worksheet = workbook.create_sheet(title=sheet_name)

                    # Convert YAML data to rows and add to worksheet
                    self._yaml_to_worksheet(yaml_data, worksheet)

                    success_count += 1

                except Exception as e:
                    print(f"Error processing file {yaml_file}: {str(e)}")
                    continue

            # If no sheets were created, create an empty one
            if success_count == 0:
                worksheet = workbook.create_sheet(title="Empty")
                worksheet.append(["No YAML files were found in the selected folder"])

            # Save the workbook
            workbook.save(file_path)

            if self.parent_window:
                color = QColor(self.styles.get('DarkTheme', {}).get('NotificationSuccess', '#6BA878'))
                self.parent_window.show_notification(
                    f"Exported {success_count}/{total_count} YAML files to Excel", color
                )

            return True

        except Exception as e:
            error_msg = f"Error exporting to Excel: {str(e)}"
            print(error_msg)
            if self.parent_window:
                color = QColor(self.styles.get('DarkTheme', {}).get('NotificationError', '#D9685A'))
                self.parent_window.show_notification(error_msg, color)
            return False

    def _find_yaml_files(self, folder_path: str) -> List[str]:
        """
        Find all YAML files in the given folder and its subfolders.
        """
        yaml_files = []

        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith(('.yaml', '.yml')):
                    yaml_files.append(os.path.join(root, file))

        return yaml_files

    def _sanitize_sheet_name(self, name: str) -> str:
        """
        Sanitize a string to be a valid Excel sheet name.
        """
        # Remove invalid characters
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            name = name.replace(char, '_')

        # Truncate to max length (Excel limit is 31 characters)
        if len(name) > 31:
            name = name[:31]

        # Ensure it's not empty
        if not name:
            name = "Sheet"

        return name

    def _yaml_to_worksheet(self, yaml_data: Any, worksheet: Worksheet):
        """
        Convert YAML data to Excel worksheet rows in a more readable format.
        For simple key-value structures, use separate rows for each key-value pair.
        For complex structures, provide both the readable format and a backup YAML string.
        """
        # Add headers
        worksheet.append(["Key", "Value", "Original YAML Type"])

        # Process the YAML data based on its structure
        self._add_yaml_data_to_worksheet(yaml_data, worksheet, "")

        # Also include the raw YAML as backup for perfect round-trip compatibility
        yaml_str = yaml.dump(yaml_data, default_flow_style=False, allow_unicode=True, sort_keys=False)
        worksheet.append(["RAW_YAML_DATA", yaml_str, "YAML_STRUCTURE"])

        # Style the header row
        self._style_header_row(worksheet)

        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet)

    def _add_yaml_data_to_worksheet(self, data: Any, worksheet: Worksheet, prefix: str = ""):
        """
        Recursively add YAML data to worksheet in a structured format for easier editing.
        """
        if isinstance(data, dict):
            for key, value in data.items():
                full_key = f"{prefix}.{key}" if prefix else key
                if isinstance(value, (dict, list)):
                    # For nested structures, add them recursively
                    self._add_yaml_data_to_worksheet(value, worksheet, full_key)
                else:
                    # Add simple key-value pairs
                    worksheet.append([full_key, str(value) if value is not None else "", type(value).__name__])
        elif isinstance(data, list):
            for i, item in enumerate(data):
                full_key = f"{prefix}[{i}]"
                if isinstance(item, (dict, list)):
                    # For nested structures in lists, add them recursively
                    self._add_yaml_data_to_worksheet(item, worksheet, full_key)
                else:
                    # Add list items with index
                    worksheet.append([full_key, str(item) if item is not None else "", type(item).__name__])
        else:
            # For scalar values at the root level
            worksheet.append([prefix if prefix else "value", str(data) if data is not None else "", type(data).__name__])

    def _style_header_row(self, worksheet: Worksheet):
        """
        Style the header row of the worksheet.
        """
        if worksheet.max_row < 1:
            return

        # Style the first row (header)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_adjust_columns(self, worksheet: Worksheet):
        """
        Auto-adjust column widths to fit content.
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width


def import_excel_dialog(parent_window):
    """
    Show dialog to import Excel file and restore the language folder structure.

    Args:
        parent_window: Main application window

    Returns:
        True if successful, False otherwise
    """
    excel_handler = ExcelHandler(parent_window)

    # Open file dialog to select Excel file
    excel_path, _ = QFileDialog.getOpenFileName(
        parent_window,
        "Import Excel File",
        "",
        "Excel Files (*.xlsx *.xls)"
    )

    if not excel_path:
        return False  # User cancelled

    # Open folder dialog to select destination
    target_folder = QFileDialog.getExistingDirectory(
        parent_window,
        "Select Target Folder for Import"
    )

    if not target_folder:
        return False  # User cancelled

    # Show progress dialog
    progress = QProgressDialog("Importing Excel file to language folder...", "Cancel", 0, 100, parent_window)
    progress.setWindowModality(Qt.WindowModal)
    progress.show()

    try:
        # Update progress
        progress.setValue(30)

        # Import from Excel to target folder
        success = excel_handler.import_from_excel(excel_path, target_folder)

        progress.setValue(80)

        if success:
            progress.setValue(100)

        progress.close()
        return success

    except Exception as e:
        progress.close()
        error_msg = f"Error during import: {str(e)}"
        if parent_window:
            color = QColor(parent_window.STYLES.get('DarkTheme', {}).get('NotificationError', '#D9685A'))
            parent_window.show_notification(error_msg, color)
        return False


def export_excel_dialog(source_folder, parent_window):
    """
    Show dialog to export language folder structure to Excel file.

    Args:
        source_folder: Source folder containing the language structure to export
        parent_window: Main application window

    Returns:
        True if successful, False otherwise
    """
    excel_handler = ExcelHandler(parent_window)

    # Open file dialog to select save location
    excel_path, _ = QFileDialog.getSaveFileName(
        parent_window,
        "Export Language Folder to Excel",
        "",
        "Excel Files (*.xlsx)"
    )

    if not excel_path:
        return False  # User cancelled

    # Ensure the file has .xlsx extension
    if not excel_path.lower().endswith('.xlsx'):
        excel_path += '.xlsx'

    # Show progress dialog
    progress = QProgressDialog("Exporting language folder to Excel...", "Cancel", 0, 100, parent_window)
    progress.setWindowModality(Qt.WindowModal)
    progress.show()

    try:
        progress.setValue(30)

        # Export language folder to Excel
        success = excel_handler.export_to_excel(source_folder, excel_path)

        progress.setValue(80)

        if success:
            progress.setValue(100)

        progress.close()
        return success

    except Exception as e:
        progress.close()
        error_msg = f"Error during export: {str(e)}"
        if parent_window:
            color = QColor(parent_window.STYLES.get('DarkTheme', {}).get('NotificationError', '#D9685A'))
            parent_window.show_notification(error_msg, color)
        return False